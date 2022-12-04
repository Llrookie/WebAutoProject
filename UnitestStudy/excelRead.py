import openpyxl as openpyxl


class excel_read:
    def read_data(self):
        # 加载工作簿
        wb = openpyxl.load_workbook('K:/PythonProject/WebAutoProject/UnitestStudy/data.xlsx')
        # 获取sheet对象
        sheet = wb['login']
        # 获取excel行数和列数
        print(sheet.max_row,sheet.max_column)
        list_all = []
        for cows in range(2,sheet.max_row+1):
            list_temp = []
            for cols in range(1,sheet.max_column+1):
                list_temp.append(sheet.cell(cows,cols).value)
            list_all.append(list_temp)
        return list_all

if __name__ == '__main__':
    excel_read().read_data()